#!/usr/bin/env python3
"""Prepare GSC Generative AI + Web page exports for deterministic triage.

Outputs:
- <prefix>_joined.csv
- <prefix>_summary.json

Supports CSV/TSV with the Python standard library. XLSX is supported when
openpyxl is available in the execution environment.

This script intentionally does NOT diagnose content quality. It performs:
- conservative URL normalization
- outer join + observation status
- export row-limit QA
- optional metadata enrichment (page type/language/business value/owner)
- cohort-aware median thresholds and percentiles
- four-quadrant assignment for matched eligible rows

Default cohort hierarchy:
  language + page_type -> page_type -> language -> global
The most specific cohort with at least --min-cohort-size eligible URLs is used.
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import statistics
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, List, Tuple
from urllib.parse import urlsplit, urlunsplit

AI_URL_ALIASES = {
    "url", "page", "pages", "网页", "页面", "排名靠前的网页", "top pages", "top page"
}
AI_IMP_ALIASES = {
    "ai impressions", "generative ai impressions", "impressions", "ai展示次数", "ai 展示次数",
    "生成式ai展示次数", "生成式 ai 展示次数", "展示次数"
}
WEB_URL_ALIASES = AI_URL_ALIASES
WEB_CLICK_ALIASES = {"clicks", "click", "点击", "点击次数"}
WEB_IMP_ALIASES = {"impressions", "impression", "展示", "展示次数", "常规展示"}
WEB_CTR_ALIASES = {"ctr", "点击率"}
WEB_POS_ALIASES = {"position", "average position", "avg position", "排名", "平均排名"}

META_URL_ALIASES = AI_URL_ALIASES
PAGE_TYPE_ALIASES = {
    "page type", "pagetype", "content type", "template", "页面类型", "网页类型", "内容类型"
}
LANGUAGE_ALIASES = {"language", "lang", "locale", "语言", "语种"}
BUSINESS_VALUE_ALIASES = {
    "business value", "commercial value", "priority value", "业务价值", "商业价值"
}
OWNER_ALIASES = {"owner", "负责人", "内容负责人", "页面负责人"}


def norm_header(value: Any) -> str:
    return " ".join(str(value or "").strip().lower().replace("_", " ").split())


def find_col(headers: List[str], aliases: set[str], label: str) -> int:
    normalized = [norm_header(h) for h in headers]
    alias_norm = {norm_header(a) for a in aliases}
    for i, h in enumerate(normalized):
        if h in alias_norm:
            return i
    raise ValueError(f"Could not identify {label} column. Headers: {headers}")


def find_optional_col(headers: List[str], aliases: set[str]) -> int | None:
    normalized = [norm_header(h) for h in headers]
    alias_norm = {norm_header(a) for a in aliases}
    for i, h in enumerate(normalized):
        if h in alias_norm:
            return i
    return None


def cell(row: List[Any], idx: int | None) -> Any:
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def to_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    s = str(value).strip().replace(",", "")
    if not s:
        return None
    if s.endswith("%"):
        try:
            return float(s[:-1]) / 100.0
        except ValueError:
            return None
    try:
        return float(s)
    except ValueError:
        return None


def normalize_url(raw: Any) -> str:
    """Conservative normalization: trim, lower scheme/host, remove fragment.

    Path case, query params, and trailing slashes are preserved.
    """
    s = str(raw or "").strip()
    if not s:
        return ""
    try:
        p = urlsplit(s)
        if not p.scheme or not p.netloc:
            return s.split("#", 1)[0]
        scheme = p.scheme.lower()
        hostname = (p.hostname or "").lower()
        port = p.port
        if port and not ((scheme == "http" and port == 80) or (scheme == "https" and port == 443)):
            netloc = f"{hostname}:{port}"
        else:
            netloc = hostname
        if p.username or p.password:
            netloc = p.netloc.lower()
        return urlunsplit((scheme, netloc, p.path, p.query, ""))
    except Exception:
        return s.split("#", 1)[0]


def read_table(path: str, sheet: str | None = None) -> Tuple[List[str], List[List[Any]]]:
    ext = Path(path).suffix.lower()
    if ext in {".csv", ".tsv", ".txt"}:
        delimiter = "\t" if ext == ".tsv" else ","
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            rows = list(csv.reader(f, delimiter=delimiter))
        if not rows:
            raise ValueError(f"Empty file: {path}")
        return rows[0], rows[1:]

    if ext in {".xlsx", ".xlsm"}:
        try:
            import openpyxl  # type: ignore
        except Exception as e:
            raise RuntimeError(
                "XLSX input requires openpyxl in this runtime. Export the relevant sheet to CSV, "
                "or use the host spreadsheet tool to convert it first."
            ) from e
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        if not rows:
            raise ValueError(f"Empty sheet in {path}")
        return [str(x or "") for x in rows[0]], rows[1:]

    raise ValueError(f"Unsupported input type: {path}")


def aggregate_ai(headers: List[str], rows: List[List[Any]]) -> Dict[str, Dict[str, Any]]:
    url_i = find_col(headers, AI_URL_ALIASES, "AI URL")
    imp_i = find_col(headers, AI_IMP_ALIASES, "AI impressions")
    agg: Dict[str, Dict[str, Any]] = {}
    for row in rows:
        url = normalize_url(cell(row, url_i))
        if not url:
            continue
        imp = to_number(cell(row, imp_i))
        if url not in agg:
            agg[url] = {"ai_impressions": 0.0, "ai_duplicate_count": 0}
        agg[url]["ai_duplicate_count"] += 1
        if imp is not None:
            agg[url]["ai_impressions"] += imp
    return agg


def aggregate_web(headers: List[str], rows: List[List[Any]]) -> Dict[str, Dict[str, Any]]:
    url_i = find_col(headers, WEB_URL_ALIASES, "Web URL")
    click_i = find_col(headers, WEB_CLICK_ALIASES, "Web clicks")
    imp_i = find_col(headers, WEB_IMP_ALIASES, "Web impressions")
    pos_i = find_col(headers, WEB_POS_ALIASES, "Web average position")

    raw: Dict[str, Dict[str, Any]] = defaultdict(lambda: {
        "web_clicks": 0.0,
        "web_impressions": 0.0,
        "position_weighted_sum": 0.0,
        "position_weight": 0.0,
        "web_duplicate_count": 0,
    })

    for row in rows:
        url = normalize_url(cell(row, url_i))
        if not url:
            continue
        clicks = to_number(cell(row, click_i)) or 0.0
        imps = to_number(cell(row, imp_i)) or 0.0
        pos = to_number(cell(row, pos_i))
        d = raw[url]
        d["web_clicks"] += clicks
        d["web_impressions"] += imps
        d["web_duplicate_count"] += 1
        if pos is not None and imps > 0:
            d["position_weighted_sum"] += pos * imps
            d["position_weight"] += imps

    out: Dict[str, Dict[str, Any]] = {}
    for url, d in raw.items():
        imps = d["web_impressions"]
        clicks = d["web_clicks"]
        out[url] = {
            "web_clicks": clicks,
            "web_impressions": imps,
            "web_ctr": clicks / imps if imps > 0 else None,
            "web_position": (
                d["position_weighted_sum"] / d["position_weight"]
                if d["position_weight"] > 0 else None
            ),
            "web_duplicate_count": d["web_duplicate_count"],
        }
    return out


def read_metadata(path: str, sheet: str | None = None) -> tuple[Dict[str, Dict[str, Any]], Dict[str, Any]]:
    headers, rows = read_table(path, sheet)
    url_i = find_col(headers, META_URL_ALIASES, "metadata URL")
    page_type_i = find_optional_col(headers, PAGE_TYPE_ALIASES)
    language_i = find_optional_col(headers, LANGUAGE_ALIASES)
    business_value_i = find_optional_col(headers, BUSINESS_VALUE_ALIASES)
    owner_i = find_optional_col(headers, OWNER_ALIASES)

    out: Dict[str, Dict[str, Any]] = {}
    duplicate_urls = 0
    for row in rows:
        url = normalize_url(cell(row, url_i))
        if not url:
            continue
        incoming = {
            "page_type": clean_text(cell(row, page_type_i)),
            "language": clean_text(cell(row, language_i)),
            "business_value": clean_text(cell(row, business_value_i)),
            "owner": clean_text(cell(row, owner_i)),
        }
        if url in out:
            duplicate_urls += 1
            # Deterministic merge: retain first non-empty value, flag conflicts.
            for k, v in incoming.items():
                if out[url].get(k) is None and v is not None:
                    out[url][k] = v
                elif v is not None and out[url].get(k) not in {None, v}:
                    out[url]["metadata_conflict"] = True
            out[url]["metadata_duplicate_count"] += 1
        else:
            out[url] = {
                **incoming,
                "metadata_duplicate_count": 1,
                "metadata_conflict": False,
            }

    info = {
        "metadata_raw_rows": len(rows),
        "metadata_unique_urls": len(out),
        "metadata_duplicate_extra_rows": duplicate_urls,
        "metadata_has_page_type": page_type_i is not None,
        "metadata_has_language": language_i is not None,
        "metadata_has_business_value": business_value_i is not None,
        "metadata_has_owner": owner_i is not None,
    }
    return out, info


def percentile_ranks(values: List[float]) -> List[float]:
    """Return average-tie percentile ranks in [0, 1]."""
    n = len(values)
    if n == 0:
        return []
    if n == 1:
        return [1.0]
    order = sorted(range(n), key=lambda i: values[i])
    ranks = [0.0] * n
    i = 0
    while i < n:
        j = i
        while j + 1 < n and values[order[j + 1]] == values[order[i]]:
            j += 1
        avg_rank = (i + j) / 2.0
        pct = avg_rank / (n - 1)
        for k in range(i, j + 1):
            ranks[order[k]] = pct
        i = j + 1
    return ranks


def norm_group_value(v: Any) -> str | None:
    s = clean_text(v)
    return s.casefold() if s else None


def pair_key(row: Dict[str, Any]) -> str | None:
    lang = norm_group_value(row.get("language"))
    pt = norm_group_value(row.get("page_type"))
    if lang and pt:
        return f"lang={lang}|page_type={pt}"
    return None


def page_type_key(row: Dict[str, Any]) -> str | None:
    pt = norm_group_value(row.get("page_type"))
    return f"page_type={pt}" if pt else None


def language_key(row: Dict[str, Any]) -> str | None:
    lang = norm_group_value(row.get("language"))
    return f"lang={lang}" if lang else None


def assign_cohorts(eligible: List[Dict[str, Any]], min_size: int) -> None:
    pair_counts: Dict[str, int] = defaultdict(int)
    type_counts: Dict[str, int] = defaultdict(int)
    lang_counts: Dict[str, int] = defaultdict(int)

    for r in eligible:
        pk = pair_key(r)
        tk = page_type_key(r)
        lk = language_key(r)
        if pk:
            pair_counts[pk] += 1
        if tk:
            type_counts[tk] += 1
        if lk:
            lang_counts[lk] += 1

    global_n = len(eligible)
    for r in eligible:
        pk = pair_key(r)
        tk = page_type_key(r)
        lk = language_key(r)
        has_pair = bool(pk)

        if pk and pair_counts[pk] >= min_size:
            level = "language+page_type"
            key = pk
            size = pair_counts[pk]
        elif tk and type_counts[tk] >= min_size:
            level = "page_type"
            key = tk
            size = type_counts[tk]
        elif lk and lang_counts[lk] >= min_size:
            level = "language"
            key = lk
            size = lang_counts[lk]
        else:
            level = "global"
            key = "global"
            size = global_n

        r["applied_cohort"] = key
        r["cohort_level"] = level
        r["cohort_size"] = size
        r["cohort_fallback_used"] = bool(has_pair and level != "language+page_type")


def assign_global_percentiles(eligible: List[Dict[str, Any]]) -> None:
    web_vals = [float(r["web_impressions"]) for r in eligible]
    ai_vals = [float(r["ai_impressions"]) for r in eligible]
    for r, wp, apct in zip(eligible, percentile_ranks(web_vals), percentile_ranks(ai_vals)):
        r["global_web_percentile"] = wp
        r["global_ai_percentile"] = apct


def assign_quadrants_by_cohort(eligible: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    groups: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for r in eligible:
        groups[r["applied_cohort"]].append(r)

    overview: Dict[str, Dict[str, Any]] = {}
    for key, rows in sorted(groups.items()):
        web_vals = [float(r["web_impressions"]) for r in rows]
        ai_vals = [float(r["ai_impressions"]) for r in rows]
        web_threshold = statistics.median(web_vals)
        ai_threshold = statistics.median(ai_vals)
        web_pcts = percentile_ranks(web_vals)
        ai_pcts = percentile_ranks(ai_vals)
        q_counts: Dict[str, int] = defaultdict(int)

        for r, wp, apct in zip(rows, web_pcts, ai_pcts):
            r["web_high_threshold"] = web_threshold
            r["ai_high_threshold"] = ai_threshold
            r["web_percentile"] = wp  # applied-cohort percentile, kept for compatibility
            r["ai_percentile"] = apct
            r["cohort_web_percentile"] = wp
            r["cohort_ai_percentile"] = apct
            wh = float(r["web_impressions"]) >= web_threshold
            ah = float(r["ai_impressions"]) >= ai_threshold
            if wh and ah:
                q = "Q1_SEARCH_HIGH_AI_HIGH"
            elif wh and not ah:
                q = "Q2_SEARCH_HIGH_AI_LOW"
            elif not wh and ah:
                q = "Q3_SEARCH_LOW_AI_HIGH"
            else:
                q = "Q4_SEARCH_LOW_AI_LOW"
            r["quadrant"] = q
            r["mismatch_intensity"] = (
                (wp - apct) if q.startswith("Q2")
                else (apct - wp) if q.startswith("Q3")
                else abs(wp - apct)
            )
            q_counts[q] += 1

        overview[key] = {
            "cohort_level": rows[0]["cohort_level"],
            "eligible_urls": len(rows),
            "web_high_threshold_median": web_threshold,
            "ai_high_threshold_median": ai_threshold,
            "quadrant_counts": dict(q_counts),
        }
    return overview


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--ai", required=True, help="Generative AI Pages export (CSV/TSV/XLSX)")
    ap.add_argument("--web", required=True, help="Web Search Pages export (CSV/TSV/XLSX)")
    ap.add_argument("--ai-sheet", default=None)
    ap.add_argument("--web-sheet", default=None)
    ap.add_argument("--metadata", default=None, help="Optional URL metadata table with page type/language/business value/owner")
    ap.add_argument("--metadata-sheet", default=None)
    ap.add_argument("--out-prefix", required=True)
    ap.add_argument("--min-web-impressions", type=float, default=0.0)
    ap.add_argument("--min-cohort-size", type=int, default=30)
    args = ap.parse_args()

    if args.min_cohort_size < 2:
        raise ValueError("--min-cohort-size must be at least 2")

    ai_headers, ai_rows = read_table(args.ai, args.ai_sheet)
    web_headers, web_rows = read_table(args.web, args.web_sheet)
    ai = aggregate_ai(ai_headers, ai_rows)
    web = aggregate_web(web_headers, web_rows)

    metadata: Dict[str, Dict[str, Any]] = {}
    metadata_info: Dict[str, Any] = {
        "metadata_raw_rows": 0,
        "metadata_unique_urls": 0,
        "metadata_duplicate_extra_rows": 0,
        "metadata_has_page_type": False,
        "metadata_has_language": False,
        "metadata_has_business_value": False,
        "metadata_has_owner": False,
    }
    if args.metadata:
        metadata, metadata_info = read_metadata(args.metadata, args.metadata_sheet)

    union = sorted(set(ai) | set(web))
    joined: List[Dict[str, Any]] = []
    for url in union:
        a = ai.get(url)
        w = web.get(url)
        m = metadata.get(url)
        if a and w:
            status = "MATCHED"
        elif a:
            status = "AI_ONLY"
        else:
            status = "WEB_ONLY"

        row: Dict[str, Any] = {
            "url": url,
            "observation_status": status,
            "ai_row_present": bool(a),
            "web_row_present": bool(w),
            "metadata_row_present": bool(m),
            "page_type": m.get("page_type") if m else None,
            "language": m.get("language") if m else None,
            "business_value": m.get("business_value") if m else None,
            "owner": m.get("owner") if m else None,
            "metadata_conflict": m.get("metadata_conflict") if m else False,
            "metadata_duplicate_count": m.get("metadata_duplicate_count") if m else 0,
            "ai_impressions": a.get("ai_impressions") if a else None,
            "web_impressions": w.get("web_impressions") if w else None,
            "web_clicks": w.get("web_clicks") if w else None,
            "web_ctr": w.get("web_ctr") if w else None,
            "web_position": w.get("web_position") if w else None,
            "ai_duplicate_count": a.get("ai_duplicate_count") if a else 0,
            "web_duplicate_count": w.get("web_duplicate_count") if w else 0,
            "ai_value_status": (
                "POSITIVE" if a and (a.get("ai_impressions") or 0) > 0
                else "REPORTED_ZERO_OR_SUPPRESSED" if a
                else "NOT_OBSERVED_IN_EXPORT"
            ),
        }
        if a and w and (w.get("web_impressions") or 0) > 0:
            row["ai_relative_exposure_index"] = a.get("ai_impressions", 0) / w["web_impressions"]
        else:
            row["ai_relative_exposure_index"] = None

        row["quadrant_eligible"] = bool(
            status == "MATCHED"
            and row["web_impressions"] is not None
            and row["ai_impressions"] is not None
            and row["web_impressions"] >= args.min_web_impressions
        )
        joined.append(row)

    eligible = [r for r in joined if r["quadrant_eligible"]]
    assign_cohorts(eligible, args.min_cohort_size)
    assign_global_percentiles(eligible)
    cohort_overview = assign_quadrants_by_cohort(eligible) if eligible else {}

    for r in joined:
        if not r["quadrant_eligible"]:
            for key in [
                "applied_cohort", "cohort_level", "cohort_size", "cohort_fallback_used",
                "web_high_threshold", "ai_high_threshold", "web_percentile", "ai_percentile",
                "cohort_web_percentile", "cohort_ai_percentile", "global_web_percentile",
                "global_ai_percentile", "quadrant", "mismatch_intensity"
            ]:
                r[key] = None

    ai_unique = len(ai)
    web_unique = len(web)
    matched = sum(r["observation_status"] == "MATCHED" for r in joined)
    ai_only = sum(r["observation_status"] == "AI_ONLY" for r in joined)
    web_only = sum(r["observation_status"] == "WEB_ONLY" for r in joined)
    union_n = len(joined)

    q_counts: Dict[str, int] = defaultdict(int)
    cohort_level_counts: Dict[str, int] = defaultdict(int)
    for r in eligible:
        q_counts[r["quadrant"]] += 1
        cohort_level_counts[r["cohort_level"]] += 1

    metadata_matched = sum(bool(r["metadata_row_present"]) for r in joined)
    page_type_covered = sum(bool(clean_text(r.get("page_type"))) for r in joined)
    language_covered = sum(bool(clean_text(r.get("language"))) for r in joined)

    warning_parts = [
        "Missing rows are not verified zeros, especially when an export hits the 1,000-row limit. "
        "Use quadrants only on MATCHED eligible rows unless you have evidence the export is exhaustive."
    ]
    if not args.metadata:
        warning_parts.append(
            "No page-type/language metadata was supplied, so all eligible URLs use a global baseline. "
            "Mixed content populations may be misclassified."
        )
    elif not (metadata_info["metadata_has_page_type"] or metadata_info["metadata_has_language"]):
        warning_parts.append(
            "Metadata was supplied but no recognized page type or language field was found; cohort-aware thresholds fell back to global."
        )

    summary = {
        "ai_export_raw_rows": len(ai_rows),
        "web_export_raw_rows": len(web_rows),
        "ai_unique_urls": ai_unique,
        "web_unique_urls": web_unique,
        "matched_urls": matched,
        "ai_only_urls": ai_only,
        "web_only_urls": web_only,
        "union_urls": union_n,
        "matched_share_of_union": matched / union_n if union_n else None,
        "ai_row_limit_risk": "HIGH" if len(ai_rows) >= 1000 else "LOWER",
        "web_row_limit_risk": "HIGH" if len(web_rows) >= 1000 else "LOWER",
        **metadata_info,
        "metadata_matched_urls": metadata_matched,
        "metadata_match_share_of_union": metadata_matched / union_n if union_n else None,
        "metadata_page_type_coverage": page_type_covered / union_n if union_n else None,
        "metadata_language_coverage": language_covered / union_n if union_n else None,
        "quadrant_eligible_urls": len(eligible),
        "min_cohort_size": args.min_cohort_size,
        "cohort_level_counts": dict(cohort_level_counts),
        "cohort_overview": cohort_overview,
        "quadrant_counts": dict(q_counts),
        "warning": " ".join(warning_parts),
    }

    prefix = Path(args.out_prefix)
    prefix.parent.mkdir(parents=True, exist_ok=True)
    csv_path = str(prefix) + "_joined.csv"
    json_path = str(prefix) + "_summary.json"

    fields = [
        "url", "observation_status", "ai_row_present", "web_row_present", "ai_value_status",
        "metadata_row_present", "page_type", "language", "business_value", "owner",
        "metadata_conflict", "metadata_duplicate_count",
        "ai_impressions", "web_impressions", "web_clicks", "web_ctr", "web_position",
        "ai_relative_exposure_index", "quadrant_eligible",
        "applied_cohort", "cohort_level", "cohort_size", "cohort_fallback_used",
        "web_high_threshold", "ai_high_threshold",
        "cohort_web_percentile", "cohort_ai_percentile",
        "global_web_percentile", "global_ai_percentile",
        "web_percentile", "ai_percentile", "quadrant", "mismatch_intensity",
        "ai_duplicate_count", "web_duplicate_count"
    ]
    with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for row in joined:
            writer.writerow({k: row.get(k) for k in fields})

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)

    print(json.dumps(summary, ensure_ascii=False, indent=2))
    print(f"Wrote: {csv_path}")
    print(f"Wrote: {json_path}")


if __name__ == "__main__":
    main()
